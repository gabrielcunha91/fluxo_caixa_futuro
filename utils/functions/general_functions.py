import streamlit as st
import pandas as pd
import numpy as np
from pandas.api.types import is_numeric_dtype
from utils.queries import *
from workalendar.america import Brazil
import openpyxl
import os
from st_aggrid import GridUpdateMode, JsCode, StAggridTheme
from st_aggrid import AgGrid, GridOptionsBuilder, JsCode, GridUpdateMode
from streamlit_echarts import st_echarts
from st_aggrid.shared import StAggridTheme
import streamlit.components.v1 as components




def export_to_excel(df, sheet_name, excel_filename):
  if os.path.exists(excel_filename):
    wb = openpyxl.load_workbook(excel_filename)
  else:
    wb = openpyxl.Workbook()

  if sheet_name in wb.sheetnames:
    wb.remove(wb[sheet_name])
  ws = wb.create_sheet(title=sheet_name)
  
  # Escrever os cabeçalhos
  for col_idx, column_title in enumerate(df.columns, start=1):
    ws.cell(row=1, column=col_idx, value=column_title)
  
  # Escrever os dados
  for row_idx, row in enumerate(df.itertuples(index=False, name=None), start=2):
    for col_idx, value in enumerate(row, start=1):
      ws.cell(row=row_idx, column=col_idx, value=value)

  wb.save(excel_filename)


def format_brazilian(num):
  try:
    num = float(num)
    return f"{num:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
  except (ValueError, TypeError):
    return num

def format_columns_brazilian(df, numeric_columns):
  for col in numeric_columns:
    if col in df.columns:
      df[col] = df[col].apply(format_brazilian)
  return df


def format_percentage(num):
  try:
    num = float(num)
    formatted_num = f"{num * 100:,.2f}"  # Multiplica por 100 e formata
    return f"{formatted_num.replace(',', 'X').replace('.', ',').replace('X', '.')}%"  # Formata como percentual
  except (ValueError, TypeError):
    return num  # Retorna o valor original em caso de erro
  
def format_columns_percentage(df, numeric_columns):
  for col in numeric_columns:
    if col in df.columns:
      df[col] = df[col].apply(format_percentage)
  return df  


def component_plotDataframe_aggrid(
    df: pd.DataFrame,
    name: str,
    num_columns: list = [],
    percent_columns: list = [],
    df_details: pd.DataFrame = None,
    coluns_merge_details: list = None,
    coluns_name_details: str = None,
    key: str = "default"
):
    """
    Exibe um DataFrame no AG Grid com:
    - filtros de conjunto em todas as colunas (agSetColumnFilter)
    - colunas numéricas com ordenação e filtro numérico (agNumberColumnFilter)
    - formatação brasileira de números e percentuais
    - colunas redimensionáveis dinamicamente para não cortar o cabeçalho
    """
    st.markdown(
        f"<h5 style='text-align: center; background-color: #ffb131; padding: 0.1em;'>{name}</h5>",
        unsafe_allow_html=True
    )

    # Formatter para números BR
    fmt_brl = JsCode(
        "function(params) { return params.value != null ? params.value.toLocaleString('pt-BR', { minimumFractionDigits: 2 }) : ''; }"
    )
    # Formatter para percentuais
    fmt_pct = JsCode(
        "function(params) { return params.value != null ? (params.value*100).toFixed(2).replace('.',',') + '%' : ''; }"
    )

    # Cria builder
    gb = GridOptionsBuilder.from_dataframe(df)
    # Configurações padrão para todas as colunas
    gb.configure_default_column(
        sortable=True,
        filter="agSetColumnFilter",
        resizable=True,
        editable=False,
        flex=1,
        minWidth=120,
        wrapHeaderText=True,
        autoHeaderHeight=True
    )

    # Converte e configura colunas numéricas
    for col in num_columns:
        if col in df.columns:
            series = df[col]
            # converte para float
            if not is_numeric_dtype(series):
                cleaned = series.astype(str).str.replace(r'[^\d,.-]', '', regex=True)
                cleaned = cleaned.str.replace(r'\.(?=\d{3},)', '', regex=True)
                numeric = cleaned.str.replace(',', '.', regex=False)
                df[col] = pd.to_numeric(numeric, errors='coerce')
            else:
                df[col] = series.astype(float)
            # configura coluna como numérica
            gb.configure_column(
                field=col,
                type=["numericColumn"],
                filter="agNumberColumnFilter",
                valueFormatter=fmt_brl
            )

    # Converte e configura colunas percentuais
    for col in percent_columns:
        if col in df.columns:
            series = df[col]
            if not is_numeric_dtype(series):
                cleaned = (
                    series.astype(str)
                          .str.replace('%', '', regex=False)
                          .str.replace(',', '.', regex=False)
                          .str.replace(r'[^\d.-]', '', regex=True)
                )
                df[col] = pd.to_numeric(cleaned, errors='coerce') / 100.0
            else:
                df[col] = series.astype(float)
            gb.configure_column(
                field=col,
                type=["numericColumn"],
                filter="agNumberColumnFilter",
                valueFormatter=fmt_pct
            )

    # Build options
    grid_options = gb.build()
    # Master-detail (opcional)
    if df_details is not None and coluns_merge_details and coluns_name_details:
        df['detail'] = df[coluns_merge_details].apply(
            lambda keys: df_details[df_details[coluns_merge_details] == keys].to_dict('records')
        )
        grid_options.update({
            "masterDetail": True,
            "columnDefs": [{"field": coluns_name_details, "cellRenderer": "agGroupCellRenderer"}] +
                           [{"field": c} for c in df.columns if c not in [coluns_name_details, 'detail']],
            "detailCellRendererParams": {
                "detailGridOptions": {"columnDefs": [{"field": c} for c in df_details.columns]},
                "getDetailRowData": JsCode("function(params){params.successCallback(params.data.detail);}")
            }
        })

    # Exibe grid
    df_to_show = df.copy()
    if 'detail' in df_to_show:
        df_to_show = df_to_show.drop(columns=['detail'])
    grid_response = AgGrid(
        df_to_show,
        gridOptions=grid_options,
        enable_enterprise_modules=True,
        update_mode=GridUpdateMode.MODEL_CHANGED,
        fit_columns_on_grid_load=True,
        allow_unsafe_jscode=True,
        key=f"aggrid_{name}_{key}",
        theme=StAggridTheme(base="balham").withParams().withParts('colorSchemeDark')
    )
    return pd.DataFrame(grid_response['data'])

    

def function_copy_dataframe_as_tsv(df: pd.DataFrame):
    # Cria uma cópia do DataFrame para não modificar o original
    df_copy = df.copy()
    
    # Identifica colunas numéricas
    numeric_columns = df_copy.select_dtypes(include=[np.number]).columns
    
    # Formata colunas numéricas substituindo "." por ","
    for col in numeric_columns:
        df_copy[col] = df_copy[col].astype(str).str.replace('.', ',', regex=False)
    
    # Gera a string TSV
    tsv = df_copy.to_csv(index=False, sep='\t')

    # HTML+JS para ocultar o textarea, copiar ao clicar e avisar o usuário
    components.html(f"""
    <textarea id="clipboard-textarea" style="position: absolute; left: -10000px;">
{tsv}
    </textarea>
    <button onclick="
        const ta = document.getElementById('clipboard-textarea');
        ta.select();
        document.execCommand('copy');
        alert('TSV copiado para a área de transferência!');
    " style="
        background-color:#1e1e1e;
        color:#fff;
        border:1px solid #333;
        padding:8px 16px;
        border-radius:4px;
        cursor:pointer;
        margin-top:8px;
    ">Copiar como TSV</button>
    """, height=80)


    